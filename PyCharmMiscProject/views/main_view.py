# views/main_view.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from models.component_db import ComponentDatabase
from models.electrical_lib import ElectricalLibrary
from models.project_manager import ProjectManager
from dialogs.component_lib_dialog import ComponentLibraryDialog
from dialogs.electrical_lib_dialog import ElectricalLibraryDialog
from dialogs.rack_add_dialog import AddRackDialog
from dialogs.add_to_rack_dialog import AddToRackDialog


class MainView:
    def __init__(self, root):
        self.root = root
        self.root.title("电气BOM管理系统")
        self.root.geometry("1400x800")
        self.root.minsize(1200, 600)

        self.component_db = ComponentDatabase()
        self.electrical_lib = ElectricalLibrary()
        self.project_mgr = ProjectManager()

        self.setup_ui()
        self.refresh_component_list()
        self.refresh_rack_display()
        self.filter_elements()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        project_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="项目", menu=project_menu)
        project_menu.add_command(label="新建项目", command=self.new_project)
        project_menu.add_command(label="打开项目", command=self.open_project)
        project_menu.add_command(label="保存项目", command=self.save_project)
        project_menu.add_separator()
        project_menu.add_command(label="导出Excel", command=self.export_excel)

        lib_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="库管理", menu=lib_menu)
        lib_menu.add_command(label="组件库管理", command=self.manage_component_lib)
        lib_menu.add_command(label="电气元件库管理", command=self.manage_electrical_lib)

        # 项目信息
        info_frame = ttk.LabelFrame(main_frame, text="项目信息", padding="5")
        info_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(info_frame, text="项目名称:").pack(side=tk.LEFT)
        self.project_name_var = tk.StringVar()
        ttk.Label(info_frame, textvariable=self.project_name_var).pack(side=tk.LEFT, padx=(10, 0))

        # 使用 PanedWindow 可调整区域大小
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        # 左侧：可用资源
        left_frame = ttk.LabelFrame(paned, text="可用资源", padding="5")
        paned.add(left_frame, weight=1)

        notebook = ttk.Notebook(left_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # 组件选项卡
        comp_tab = ttk.Frame(notebook)
        notebook.add(comp_tab, text="组件")
        search_frame = ttk.Frame(comp_tab)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        self.comp_search_var = tk.StringVar()
        self.comp_search_var.trace('w', self.filter_components)
        ttk.Entry(search_frame, textvariable=self.comp_search_var, width=20).pack(side=tk.LEFT, padx=(5, 0))

        self.comp_tree = ttk.Treeview(comp_tab, columns=("name",), show="headings", height=20)
        self.comp_tree.heading("name", text="组件名称")
        self.comp_tree.column("name", width=200)
        scroll_comp = ttk.Scrollbar(comp_tab, orient=tk.VERTICAL, command=self.comp_tree.yview)
        self.comp_tree.configure(yscrollcommand=scroll_comp.set)
        self.comp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_comp.pack(side=tk.RIGHT, fill=tk.Y)

        # 元件选项卡
        elem_tab = ttk.Frame(notebook)
        notebook.add(elem_tab, text="电气元件")
        search_elem_frame = ttk.Frame(elem_tab)
        search_elem_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(search_elem_frame, text="搜索:").pack(side=tk.LEFT)
        self.elem_search_var = tk.StringVar()
        self.elem_search_var.trace('w', self.filter_elements)
        ttk.Entry(search_elem_frame, textvariable=self.elem_search_var, width=20).pack(side=tk.LEFT, padx=(5, 0))

        self.elem_tree = ttk.Treeview(elem_tab, columns=("part_number", "name", "specification"), show="headings",
                                      height=20)
        self.elem_tree.heading("part_number", text="物料编码")
        self.elem_tree.heading("name", text="物料名称")
        self.elem_tree.heading("specification", text="规格")
        self.elem_tree.column("part_number", width=120)
        self.elem_tree.column("name", width=150)
        self.elem_tree.column("specification", width=150)

        scroll_elem = ttk.Scrollbar(elem_tab, orient=tk.VERTICAL, command=self.elem_tree.yview)
        self.elem_tree.configure(yscrollcommand=scroll_elem.set)
        self.elem_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_elem.pack(side=tk.RIGHT, fill=tk.Y)

        # 中间：机架列表 + 竖向按钮
        middle_frame = ttk.LabelFrame(paned, text="机架列表", padding="5")
        paned.add(middle_frame, weight=1)

        # 机架树
        self.rack_tree = ttk.Treeview(middle_frame, columns=("name",), show="headings", height=20)
        self.rack_tree.heading("name", text="机架名称")
        self.rack_tree.column("name", width=150)
        rack_scroll = ttk.Scrollbar(middle_frame, orient=tk.VERTICAL, command=self.rack_tree.yview)
        self.rack_tree.configure(yscrollcommand=rack_scroll.set)
        self.rack_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rack_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 机架操作按钮（竖向）
        rack_btn_frame = ttk.Frame(middle_frame)
        rack_btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        ttk.Button(rack_btn_frame, text="新建机架", command=self.add_rack).pack(fill=tk.X, pady=2)
        ttk.Button(rack_btn_frame, text="重命名", command=self.rename_rack).pack(fill=tk.X, pady=2)
        ttk.Button(rack_btn_frame, text="删除机架", command=self.del_rack).pack(fill=tk.X, pady=2)

        # 右侧：机架内容 + 竖向按钮
        right_frame = ttk.LabelFrame(paned, text="机架内容", padding="5")
        paned.add(right_frame, weight=2)

        self.rack_items_tree = ttk.Treeview(right_frame, columns=("type", "name", "quantity"), show="headings")
        self.rack_items_tree.heading("type", text="类型")
        self.rack_items_tree.heading("name", text="名称")
        self.rack_items_tree.heading("quantity", text="数量")
        self.rack_items_tree.column("type", width=80)
        self.rack_items_tree.column("name", width=250)
        self.rack_items_tree.column("quantity", width=80)
        items_scroll = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.rack_items_tree.yview)
        self.rack_items_tree.configure(yscrollcommand=items_scroll.set)
        self.rack_items_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        items_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 内容操作按钮（竖向）
        items_btn_frame = ttk.Frame(right_frame)
        items_btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        ttk.Button(items_btn_frame, text="添加选中到当前机架", command=self.add_selected_to_rack).pack(fill=tk.X, pady=2)
        ttk.Button(items_btn_frame, text="删除选中条目", command=self.del_rack_item).pack(fill=tk.X, pady=2)
        ttk.Button(items_btn_frame, text="修改数量", command=self.edit_rack_item_qty).pack(fill=tk.X, pady=2)

        # 底部按钮
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=10)
        ttk.Button(bottom_frame, text="生成BOM", command=self.generate_bom).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_frame, text="导出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)

        # 绑定事件
        self.rack_tree.bind('<<TreeviewSelect>>', self.on_rack_select)
        self.comp_tree.bind('<Double-1>', lambda e: self.add_selected_to_rack())
        self.elem_tree.bind('<Double-1>', lambda e: self.add_selected_to_rack())

    def refresh_component_list(self):
        self.comp_tree.delete(*self.comp_tree.get_children())
        search = self.comp_search_var.get()
        comps = self.component_db.get_all_components(search_term=search)
        for c in comps:
            self.comp_tree.insert("", tk.END, values=(c["name"],))

    def filter_components(self, *args):
        self.refresh_component_list()

    def filter_elements(self, *args):
        self.elem_tree.delete(*self.elem_tree.get_children())
        search = self.elem_search_var.get()
        elements = self.electrical_lib.get_all_elements(search_term=search, include_replaced=True)
        for e in elements:
            self.elem_tree.insert("", tk.END, values=(e["part_number"], e["name"], e["specification"]))

    def refresh_rack_display(self):
        self.rack_tree.delete(*self.rack_tree.get_children())
        for rack in self.project_mgr.current_project["racks"]:
            self.rack_tree.insert("", tk.END, values=(rack["name"],))
        self.rack_items_tree.delete(*self.rack_items_tree.get_children())

    def on_rack_select(self, event):
        selection = self.rack_tree.selection()
        if not selection:
            return
        index = self.rack_tree.index(selection[0])
        rack = self.project_mgr.current_project["racks"][index]
        self.rack_items_tree.delete(*self.rack_items_tree.get_children())
        for item in rack["items"]:
            type_text = "组件" if item["type"] == "component" else "元件"
            name_text = item["id"]
            self.rack_items_tree.insert("", tk.END, values=(type_text, name_text, item["quantity"]))

    def add_rack(self):
        dialog = AddRackDialog(self.root)
        name = dialog.show()
        if name:
            self.project_mgr.add_rack(name)
            self.refresh_rack_display()

    def rename_rack(self):
        selection = self.rack_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要重命名的机架")
            return
        index = self.rack_tree.index(selection[0])
        old_name = self.project_mgr.current_project["racks"][index]["name"]
        new_name = self.simple_input_dialog(self.root, "重命名机架", f"请输入新名称 (原: {old_name})")
        if new_name:
            self.project_mgr.rename_rack(index, new_name)
            self.refresh_rack_display()

    def del_rack(self):
        selection = self.rack_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的机架")
            return
        if messagebox.askyesno("确认", "删除机架会同时删除其中的所有条目，确定吗？"):
            index = self.rack_tree.index(selection[0])
            self.project_mgr.remove_rack(index)
            self.refresh_rack_display()

    def add_selected_to_rack(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            messagebox.showwarning("警告", "请先选择一个机架")
            return
        rack_index = self.rack_tree.index(rack_sel[0])

        # 判断焦点在组件还是元件
        focused = self.root.focus_get()
        if focused in [self.comp_tree, self.comp_tree.get_children()]:
            comp_sel = self.comp_tree.selection()
            if not comp_sel:
                messagebox.showwarning("警告", "请选择要添加的组件")
                return
            comp_name = self.comp_tree.item(comp_sel[0])["values"][0]
            dialog = AddToRackDialog(self.root, "component", comp_name, comp_name)
            qty = dialog.show()
            if qty:
                self.project_mgr.add_item_to_rack(rack_index, "component", comp_name, qty)
                self.on_rack_select(None)
        else:
            elem_sel = self.elem_tree.selection()
            if not elem_sel:
                messagebox.showwarning("警告", "请选择要添加的电气元件")
                return
            values = self.elem_tree.item(elem_sel[0])["values"]
            part_number = values[0]
            name = values[1]
            dialog = AddToRackDialog(self.root, "element", f"{name} ({part_number})", part_number)
            qty = dialog.show()
            if qty:
                self.project_mgr.add_item_to_rack(rack_index, "element", part_number, qty)
                self.on_rack_select(None)

    def del_rack_item(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        item_sel = self.rack_items_tree.selection()
        if not item_sel:
            messagebox.showwarning("警告", "请选择要删除的条目")
            return
        item_index = self.rack_items_tree.index(item_sel[0])
        if messagebox.askyesno("确认", "确定删除该条目吗？"):
            self.project_mgr.remove_item_from_rack(rack_index, item_index)
            self.on_rack_select(None)

    def edit_rack_item_qty(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        item_sel = self.rack_items_tree.selection()
        if not item_sel:
            messagebox.showwarning("警告", "请选择要修改数量的条目")
            return
        item_index = self.rack_items_tree.index(item_sel[0])
        old_qty = self.project_mgr.current_project["racks"][rack_index]["items"][item_index]["quantity"]
        new_qty_str = self.simple_input_dialog(self.root, "修改数量", f"请输入新数量 (原: {old_qty})")
        if new_qty_str:
            try:
                new_qty = int(new_qty_str)
                if new_qty <= 0:
                    raise ValueError
                self.project_mgr.update_item_quantity(rack_index, item_index, new_qty)
                self.on_rack_select(None)
            except:
                messagebox.showerror("错误", "数量必须是正整数")

    def new_project(self):
        name = self.simple_input_dialog(self.root, "新建项目", "项目名称:")
        if name:
            self.project_mgr.new_project(name)
            self.project_name_var.set(name)
            self.refresh_rack_display()

    def open_project(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if file_path:
            self.project_mgr.load_project(file_path)
            self.project_name_var.set(self.project_mgr.current_project["name"])
            self.refresh_rack_display()

    def save_project(self):
        if not self.project_mgr.file_path:
            file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
            if not file_path:
                return
            self.project_mgr.save_project(file_path)
        else:
            self.project_mgr.save_project()
        messagebox.showinfo("成功", "项目保存成功")

    def manage_component_lib(self):
        dialog = ComponentLibraryDialog(self.root, self.component_db)
        self.root.wait_window(dialog.dialog)
        self.refresh_component_list()

    def manage_electrical_lib(self):
        dialog = ElectricalLibraryDialog(self.root, self.electrical_lib, quantity_mode=False)
        self.root.wait_window(dialog.dialog)
        self.filter_elements()

    def generate_bom(self):
        if not self.project_mgr.current_project["racks"]:
            messagebox.showwarning("警告", "项目中还没有任何机架，请先添加机架和条目")
            return

        bom_window = tk.Toplevel(self.root)
        bom_window.title("电气BOM详情")
        bom_window.geometry("1200x600")
        bom_window.transient(self.root)
        bom_window.grab_set()

        main_frame = ttk.Frame(bom_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text=f"电气BOM - {self.project_mgr.current_project['name']}",
                  font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(0, 10))

        # 去除“所属组件”列
        columns = ("rack", "placement", "material", "part_number", "specification", "quantity", "notes")
        tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=25)
        tree.heading("rack", text="机架")
        tree.heading("placement", text="位置")
        tree.heading("material", text="材料名称")
        tree.heading("part_number", text="料号")
        tree.heading("specification", text="规格")
        tree.heading("quantity", text="数量")
        tree.heading("notes", text="备注")
        for col in columns:
            tree.column(col, width=120)
        tree.column("material", width=150)

        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 收集BOM数据（用于导出）
        bom_data = []

        for rack in self.project_mgr.current_project["racks"]:
            rack_name = rack["name"]
            for item in rack["items"]:
                if item["type"] == "element":
                    elem = self.electrical_lib.get_element(item["id"])
                    if elem:
                        final_pn = self.electrical_lib.get_replacement_chain(item["id"]) or item["id"]
                        final_elem = self.electrical_lib.get_element(final_pn) or elem
                        # 元件没有预制板/零散件区分，统一放在“零散件”位置？用户要求元件也区分？根据需求，元件可视为零散件，这里标记为“零散件”
                        placement_cn = "零散件"
                        tree.insert("", tk.END, values=(
                            rack_name, placement_cn,
                            final_elem["name"], final_pn, final_elem["specification"],
                            item["quantity"], ""
                        ))
                        bom_data.append((rack_name, placement_cn, final_elem["name"], final_pn,
                                         final_elem["specification"], item["quantity"], ""))
                else:  # component
                    comp = self.component_db.get_component(item["id"])
                    if comp:
                        for part in comp["electrical_parts"]:
                            # 预制板数量
                            if part["prefab_qty"] > 0:
                                total_prefab = part["prefab_qty"] * item["quantity"]
                                final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                final_elem = self.electrical_lib.get_element(final_pn)
                                material_name = final_elem["name"] if final_elem else part["material"]
                                spec = final_elem["specification"] if final_elem else part["specification"]
                                tree.insert("", tk.END, values=(
                                    rack_name, "预制板", material_name, final_pn, spec, total_prefab, part["notes"]
                                ))
                                bom_data.append((rack_name, "预制板", material_name, final_pn, spec, total_prefab, part["notes"]))
                            # 零散件数量
                            if part["spare_qty"] > 0:
                                total_spare = part["spare_qty"] * item["quantity"]
                                final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                final_elem = self.electrical_lib.get_element(final_pn)
                                material_name = final_elem["name"] if final_elem else part["material"]
                                spec = final_elem["specification"] if final_elem else part["specification"]
                                tree.insert("", tk.END, values=(
                                    rack_name, "零散件", material_name, final_pn, spec, total_spare, part["notes"]
                                ))
                                bom_data.append((rack_name, "零散件", material_name, final_pn, spec, total_spare, part["notes"]))

        # 导出按钮
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="导出Excel", command=lambda: self.export_bom_excel(bom_data)).pack()

    def export_bom_excel(self, bom_data):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        try:
            # 1. 处理物料明细：合并相同机架+位置+料号的数量
            # bom_data 格式：[(机架, 位置, 材料名称, 料号, 规格, 数量, 备注), ...]
            # 先转换为DataFrame
            df = pd.DataFrame(bom_data, columns=["机架", "位置", "材料名称", "料号", "规格", "数量", "备注"])

            # 按机架、位置、料号、材料名称、规格、备注分组，合并数量
            grouped = df.groupby(["机架", "位置", "料号", "材料名称", "规格", "备注"], as_index=False)["数量"].sum()

            # 按机架、位置排序（确保预制板在前，零散件在后？可按位置自然顺序，但最好预制板在前）
            # 定义位置顺序
            position_order = {"预制板": 0, "零散件": 1}
            grouped["位置排序"] = grouped["位置"].map(position_order)
            grouped = grouped.sort_values(["机架", "位置排序", "料号"]).drop(columns=["位置排序"])

            # 重新排列列顺序
            df_material = grouped[["机架", "位置", "料号", "材料名称", "规格", "数量", "备注"]]

            # 2. 写入物料明细到Excel（先不合并单元格）
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_material.to_excel(writer, sheet_name="物料明细", index=False)

            # 3. 合并单元格（相同机架、相同位置的行合并）
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            wb = load_workbook(file_path)
            ws = wb["物料明细"]

            # 记录需要合并的起止行
            merge_map = {}  # key: (机架, 位置) -> [start_row, end_row]
            current_key = None
            start_row = 2
            for row in range(2, ws.max_row + 2):  # 多一行用于结束
                rack = ws.cell(row=row, column=1).value if row <= ws.max_row else None
                pos = ws.cell(row=row, column=2).value if row <= ws.max_row else None
                key = (rack, pos)
                if key != current_key:
                    if current_key is not None:
                        merge_map[current_key] = (start_row, row - 1)
                    current_key = key
                    start_row = row

            # 执行合并
            for (rack, pos), (s, e) in merge_map.items():
                if s < e:
                    ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                    ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)
                    # 居中
                    ws.cell(row=s, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=s, column=2).alignment = Alignment(horizontal='center', vertical='center')

            wb.save(file_path)

            # 4. 收集IO点并生成IO分配表
            io_rows = []
            for rack in self.project_mgr.current_project["racks"]:
                rack_name = rack["name"]
                input_names = []
                output_names = []
                for item in rack["items"]:
                    if item["type"] != "component":
                        continue
                    comp = self.component_db.get_component(item["id"])
                    if not comp or not comp.get("io_points"):
                        continue
                    io_points = comp["io_points"]
                    for instance in range(1, item["quantity"] + 1):
                        for name in io_points.get("inputs", []):
                            full_name = f"{comp['name']}{instance}{name}"
                            input_names.append(full_name)
                        for name in io_points.get("outputs", []):
                            full_name = f"{comp['name']}{instance}{name}"
                            output_names.append(full_name)
                for idx, name in enumerate(input_names, start=1):
                    io_rows.append({"机架": rack_name, "点类型": "输入", "序号": idx, "点名称": name})
                for idx, name in enumerate(output_names, start=1):
                    io_rows.append({"机架": rack_name, "点类型": "输出", "序号": idx, "点名称": name})
            df_io = pd.DataFrame(io_rows) if io_rows else pd.DataFrame(columns=["机架", "点类型", "序号", "点名称"])
            # 将IO表追加到同一个Excel文件
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_io.to_excel(writer, sheet_name="IO分配表", index=False)

            messagebox.showinfo("成功", f"BOM已导出到: {file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def export_excel(self):
        if self.project_mgr.current_project["racks"]:
            # 临时生成BOM数据
            bom_data = []
            for rack in self.project_mgr.current_project["racks"]:
                rack_name = rack["name"]
                for item in rack["items"]:
                    if item["type"] == "element":
                        elem = self.electrical_lib.get_element(item["id"])
                        if elem:
                            final_pn = self.electrical_lib.get_replacement_chain(item["id"]) or item["id"]
                            final_elem = self.electrical_lib.get_element(final_pn) or elem
                            bom_data.append((rack_name, "零散件", final_elem["name"], final_pn,
                                             final_elem["specification"], item["quantity"], ""))
                    else:
                        comp = self.component_db.get_component(item["id"])
                        if comp:
                            for part in comp["electrical_parts"]:
                                if part["prefab_qty"] > 0:
                                    total = part["prefab_qty"] * item["quantity"]
                                    final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                    final_elem = self.electrical_lib.get_element(final_pn)
                                    name = final_elem["name"] if final_elem else part["material"]
                                    spec = final_elem["specification"] if final_elem else part["specification"]
                                    bom_data.append((rack_name, "预制板", name, final_pn, spec, total, part["notes"]))
                                if part["spare_qty"] > 0:
                                    total = part["spare_qty"] * item["quantity"]
                                    final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                    final_elem = self.electrical_lib.get_element(final_pn)
                                    name = final_elem["name"] if final_elem else part["material"]
                                    spec = final_elem["specification"] if final_elem else part["specification"]
                                    bom_data.append((rack_name, "零散件", name, final_pn, spec, total, part["notes"]))
            self.export_bom_excel(bom_data)
        else:
            messagebox.showwarning("警告", "项目为空，无法导出")

    def simple_input_dialog(self, parent, title, prompt):
        """简单的输入对话框"""
        result = [None]
        dialog = tk.Toplevel(parent)
        dialog.title(title)
        dialog.geometry("300x120")
        dialog.transient(parent)
        dialog.grab_set()
        # 居中
        dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=prompt).pack(anchor=tk.W, pady=(0, 5))
        entry = ttk.Entry(frame, width=30)
        entry.pack(fill=tk.X, pady=(0, 10))
        entry.focus()
        def ok():
            result[0] = entry.get().strip()
            dialog.destroy()
        def cancel():
            result[0] = None
            dialog.destroy()
        btn_frame = ttk.Frame(frame)
        btn_frame.pack()
        ttk.Button(btn_frame, text="确定", command=ok).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="取消", command=cancel).pack(side=tk.LEFT)
        entry.bind('<Return>', lambda e: ok())
        dialog.wait_window()
        return result[0]