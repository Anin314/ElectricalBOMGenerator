# views/main_view.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from models.component_db import ComponentDatabase
from models.electrical_lib import ElectricalLibrary
from models.project_manager import ProjectManager
from dialogs.component_lib_dialog import ComponentLibraryDialog
from dialogs.electrical_lib_dialog import ElectricalLibraryDialog
from dialogs.rack_add_dialog import AddRackDialog
from dialogs.add_to_rack_dialog import AddToRackDialog


class MainView:
    def __init__(self, root):
        self.root = root
        self.root.title("电气BOM管理系统")
        self.root.geometry("1400x800")
        self.root.minsize(1200, 600)

        self.component_db = ComponentDatabase()
        self.electrical_lib = ElectricalLibrary()
        self.project_mgr = ProjectManager()

        self.setup_ui()
        self.refresh_component_list()
        self.refresh_rack_display()
        self.filter_elements()

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        project_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="项目", menu=project_menu)
        project_menu.add_command(label="新建项目", command=self.new_project)
        project_menu.add_command(label="打开项目", command=self.open_project)
        project_menu.add_command(label="保存项目", command=self.save_project)
        project_menu.add_separator()
        project_menu.add_command(label="导出Excel", command=self.export_excel)

        lib_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="库管理", menu=lib_menu)
        lib_menu.add_command(label="组件库管理", command=self.manage_component_lib)
        lib_menu.add_command(label="电气元件库管理", command=self.manage_electrical_lib)

        # 项目信息
        info_frame = ttk.LabelFrame(main_frame, text="项目信息", padding="5")
        info_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(info_frame, text="项目名称:").pack(side=tk.LEFT)
        self.project_name_var = tk.StringVar()
        ttk.Label(info_frame, textvariable=self.project_name_var).pack(side=tk.LEFT, padx=(10, 0))

        # 使用 PanedWindow 可调整区域大小
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        # 左侧：可用资源
        left_frame = ttk.LabelFrame(paned, text="可用资源", padding="5")
        paned.add(left_frame, weight=1)

        notebook = ttk.Notebook(left_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # 组件选项卡
        comp_tab = ttk.Frame(notebook)
        notebook.add(comp_tab, text="组件")
        search_frame = ttk.Frame(comp_tab)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        self.comp_search_var = tk.StringVar()
        self.comp_search_var.trace('w', self.filter_components)
        ttk.Entry(search_frame, textvariable=self.comp_search_var, width=20).pack(side=tk.LEFT, padx=(5, 0))

        self.comp_tree = ttk.Treeview(comp_tab, columns=("name",), show="headings", height=20)
        self.comp_tree.heading("name", text="组件名称")
        self.comp_tree.column("name", width=200)
        scroll_comp = ttk.Scrollbar(comp_tab, orient=tk.VERTICAL, command=self.comp_tree.yview)
        self.comp_tree.configure(yscrollcommand=scroll_comp.set)
        self.comp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_comp.pack(side=tk.RIGHT, fill=tk.Y)

        # 元件选项卡
        elem_tab = ttk.Frame(notebook)
        notebook.add(elem_tab, text="电气元件")
        search_elem_frame = ttk.Frame(elem_tab)
        search_elem_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Label(search_elem_frame, text="搜索:").pack(side=tk.LEFT)
        self.elem_search_var = tk.StringVar()
        self.elem_search_var.trace('w', self.filter_elements)
        ttk.Entry(search_elem_frame, textvariable=self.elem_search_var, width=20).pack(side=tk.LEFT, padx=(5, 0))

        self.elem_tree = ttk.Treeview(elem_tab, columns=("part_number", "name", "specification"), show="headings",
                                      height=20)
        self.elem_tree.heading("part_number", text="物料编码")
        self.elem_tree.heading("name", text="物料名称")
        self.elem_tree.heading("specification", text="规格")
        self.elem_tree.column("part_number", width=120)
        self.elem_tree.column("name", width=150)
        self.elem_tree.column("specification", width=150)

        scroll_elem = ttk.Scrollbar(elem_tab, orient=tk.VERTICAL, command=self.elem_tree.yview)
        self.elem_tree.configure(yscrollcommand=scroll_elem.set)
        self.elem_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_elem.pack(side=tk.RIGHT, fill=tk.Y)

        # 中间：机架列表 + 竖向按钮
        middle_frame = ttk.LabelFrame(paned, text="机架列表", padding="5")
        paned.add(middle_frame, weight=1)

        # 机架树
        self.rack_tree = ttk.Treeview(middle_frame, columns=("name",), show="headings", height=20)
        self.rack_tree.heading("name", text="机架名称")
        self.rack_tree.column("name", width=150)
        rack_scroll = ttk.Scrollbar(middle_frame, orient=tk.VERTICAL, command=self.rack_tree.yview)
        self.rack_tree.configure(yscrollcommand=rack_scroll.set)
        self.rack_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rack_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 机架操作按钮（竖向）
        rack_btn_frame = ttk.Frame(middle_frame)
        rack_btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        ttk.Button(rack_btn_frame, text="新建机架", command=self.add_rack).pack(fill=tk.X, pady=2)
        ttk.Button(rack_btn_frame, text="重命名", command=self.rename_rack).pack(fill=tk.X, pady=2)
        ttk.Button(rack_btn_frame, text="删除机架", command=self.del_rack).pack(fill=tk.X, pady=2)

        # 右侧：机架内容 + 机架物料明细
        right_frame = ttk.LabelFrame(paned, text="机架详情", padding="5")
        paned.add(right_frame, weight=2)

        # 使用垂直分割的 PanedWindow
        right_paned = ttk.PanedWindow(right_frame, orient=tk.VERTICAL)
        right_paned.pack(fill=tk.BOTH, expand=True)

        # 上部：机架内容列表（高度缩小）
        top_frame = ttk.Frame(right_paned)
        right_paned.add(top_frame, weight=1)

        self.rack_items_tree = ttk.Treeview(top_frame, columns=("type", "name", "quantity"), show="headings")
        self.rack_items_tree.heading("type", text="类型")
        self.rack_items_tree.heading("name", text="名称")
        self.rack_items_tree.heading("quantity", text="数量")
        self.rack_items_tree.column("type", width=80)
        self.rack_items_tree.column("name", width=250)
        self.rack_items_tree.column("quantity", width=80)
        items_scroll = ttk.Scrollbar(top_frame, orient=tk.VERTICAL, command=self.rack_items_tree.yview)
        self.rack_items_tree.configure(yscrollcommand=items_scroll.set)
        self.rack_items_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        items_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 操作按钮（竖向）放在右侧，为了节省空间，将按钮移到机架内容列表的右侧（原有位置）
        items_btn_frame = ttk.Frame(top_frame)
        items_btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        ttk.Button(items_btn_frame, text="添加选中到当前机架", command=self.add_selected_to_rack).pack(fill=tk.X,
                                                                                                       pady=2)
        ttk.Button(items_btn_frame, text="删除选中条目", command=self.del_rack_item).pack(fill=tk.X, pady=2)
        ttk.Button(items_btn_frame, text="修改数量", command=self.edit_rack_item_qty).pack(fill=tk.X, pady=2)

        # 下部：机架电气元件明细列表（只显示直接添加的电气元件）
        bottom_frame = ttk.LabelFrame(right_paned, text="机架电气元件明细", padding="5")
        right_paned.add(bottom_frame, weight=1)

        # 明细区域使用水平容器
        bottom_container = ttk.Frame(bottom_frame)
        bottom_container.pack(fill=tk.BOTH, expand=True)

        columns_detail = ("part_number", "name", "specification", "quantity", "source")
        self.rack_detail_tree = ttk.Treeview(bottom_container, columns=columns_detail, show="headings")
        self.rack_detail_tree.heading("part_number", text="料号")
        self.rack_detail_tree.heading("name", text="名称")
        self.rack_detail_tree.heading("specification", text="规格")
        self.rack_detail_tree.heading("quantity", text="数量")
        self.rack_detail_tree.heading("source", text="来源")
        # 调整列宽（适当减小，留出按钮空间）
        self.rack_detail_tree.column("part_number", width=90)
        self.rack_detail_tree.column("name", width=120)
        self.rack_detail_tree.column("specification", width=120)
        self.rack_detail_tree.column("quantity", width=60)
        self.rack_detail_tree.column("source", width=80)

        detail_scroll = ttk.Scrollbar(bottom_container, orient=tk.VERTICAL, command=self.rack_detail_tree.yview)
        self.rack_detail_tree.configure(yscrollcommand=detail_scroll.set)
        self.rack_detail_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 明细列表的操作按钮（竖向）
        detail_btn_frame = ttk.Frame(bottom_container)
        detail_btn_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        ttk.Button(detail_btn_frame, text="删除选中", command=self.del_rack_detail_item).pack(fill=tk.X, pady=2)
        ttk.Button(detail_btn_frame, text="修改数量", command=self.edit_rack_detail_qty).pack(fill=tk.X, pady=2)

        # 底部按钮
        bottom_frame_main = ttk.Frame(main_frame)
        bottom_frame_main.pack(fill=tk.X, pady=10)
        ttk.Button(bottom_frame_main, text="生成BOM", command=self.generate_bom).pack(side=tk.LEFT, padx=5)
        ttk.Button(bottom_frame_main, text="导出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)

        # 绑定事件
        self.rack_tree.bind('<<TreeviewSelect>>', self.on_rack_select)
        self.comp_tree.bind('<Double-1>', lambda e: self.add_selected_to_rack())
        self.elem_tree.bind('<Double-1>', lambda e: self.add_selected_to_rack())
        self.rack_items_tree.bind('<Double-1>', self.show_component_details_popup)

    def refresh_component_list(self):
        self.comp_tree.delete(*self.comp_tree.get_children())
        search = self.comp_search_var.get()
        comps = self.component_db.get_all_components(search_term=search)
        for c in comps:
            self.comp_tree.insert("", tk.END, values=(c["name"],))

    def filter_components(self, *args):
        self.refresh_component_list()

    def filter_elements(self, *args):
        self.elem_tree.delete(*self.elem_tree.get_children())
        search = self.elem_search_var.get()
        elements = self.electrical_lib.get_all_elements(search_term=search, include_replaced=True)
        for e in elements:
            self.elem_tree.insert("", tk.END, values=(e["part_number"], e["name"], e["specification"]))

    def refresh_rack_display(self):
        """刷新机架列表，并清空内容区和明细区"""
        self.rack_tree.delete(*self.rack_tree.get_children())
        for rack in self.project_mgr.current_project["racks"]:
            self.rack_tree.insert("", tk.END, values=(rack["name"],))
        self.rack_items_tree.delete(*self.rack_items_tree.get_children())
        self.rack_detail_tree.delete(*self.rack_detail_tree.get_children())

    def on_rack_select(self, event=None):
        selection = self.rack_tree.selection()
        if not selection:
            return
        index = self.rack_tree.index(selection[0])
        rack = self.project_mgr.current_project["racks"][index]
        # 更新机架内容列表：只显示组件，不显示直接添加的电气元件
        self.rack_items_tree.delete(*self.rack_items_tree.get_children())
        for item in rack["items"]:
            if item["type"] == "component":
                type_text = "组件"
                name_text = item["id"]
                self.rack_items_tree.insert("", tk.END, values=(type_text, name_text, item["quantity"]))
        # 刷新下方的明细列表（只显示直接添加的电气元件）
        self.refresh_rack_detail(rack)

    def refresh_rack_detail(self, rack):
        """根据机架内容，只显示直接添加到机架的电气元件（不包括组件内的物料）"""
        self.rack_detail_tree.delete(*self.rack_detail_tree.get_children())
        for item in rack["items"]:
            if item["type"] == "element":
                elem = self.electrical_lib.get_element(item["id"])
                if elem:
                    final_pn = self.electrical_lib.get_replacement_chain(item["id"]) or item["id"]
                    final_elem = self.electrical_lib.get_element(final_pn) or elem
                    self.rack_detail_tree.insert("", tk.END, values=(
                        final_pn,
                        final_elem["name"],
                        final_elem["specification"],
                        item["quantity"],
                        "直接添加"
                    ))
            # 组件类型不展开，忽略

    def add_rack(self):
        dialog = AddRackDialog(self.root)
        name = dialog.show()
        if name:
            self.project_mgr.add_rack(name)
            self.refresh_rack_display()

    def rename_rack(self):
        selection = self.rack_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要重命名的机架")
            return
        index = self.rack_tree.index(selection[0])
        old_name = self.project_mgr.current_project["racks"][index]["name"]
        new_name = self.simple_input_dialog(self.root, "重命名机架", f"请输入新名称 (原: {old_name})")
        if new_name:
            self.project_mgr.rename_rack(index, new_name)
            self.refresh_rack_display()

    def del_rack(self):
        selection = self.rack_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的机架")
            return
        if messagebox.askyesno("确认", "删除机架会同时删除其中的所有条目，确定吗？"):
            index = self.rack_tree.index(selection[0])
            self.project_mgr.remove_rack(index)
            self.refresh_rack_display()

    def add_selected_to_rack(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            messagebox.showwarning("警告", "请先选择一个机架")
            return
        rack_index = self.rack_tree.index(rack_sel[0])

        # 判断焦点在组件还是元件
        focused = self.root.focus_get()
        if focused in [self.comp_tree, self.comp_tree.get_children()]:
            comp_sel = self.comp_tree.selection()
            if not comp_sel:
                messagebox.showwarning("警告", "请选择要添加的组件")
                return
            comp_name = self.comp_tree.item(comp_sel[0])["values"][0]
            dialog = AddToRackDialog(self.root, "component", comp_name, comp_name)
            qty = dialog.show()
            if qty:
                self.project_mgr.add_item_to_rack(rack_index, "component", comp_name, qty)
        else:
            elem_sel = self.elem_tree.selection()
            if not elem_sel:
                messagebox.showwarning("警告", "请选择要添加的电气元件")
                return
            values = self.elem_tree.item(elem_sel[0])["values"]
            part_number = values[0]
            name = values[1]
            dialog = AddToRackDialog(self.root, "element", f"{name} ({part_number})", part_number)
            qty = dialog.show()
            if qty:
                self.project_mgr.add_item_to_rack(rack_index, "element", part_number, qty)
        # 刷新当前选中机架的显示
        self.on_rack_select()

    def del_rack_item(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        item_sel = self.rack_items_tree.selection()
        if not item_sel:
            messagebox.showwarning("警告", "请选择要删除的条目")
            return
        item_index = self.rack_items_tree.index(item_sel[0])
        if messagebox.askyesno("确认", "确定删除该条目吗？"):
            self.project_mgr.remove_item_from_rack(rack_index, item_index)
            self.on_rack_select()

    def edit_rack_item_qty(self):
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        item_sel = self.rack_items_tree.selection()
        if not item_sel:
            messagebox.showwarning("警告", "请选择要修改数量的条目")
            return
        item_index = self.rack_items_tree.index(item_sel[0])
        old_qty = self.project_mgr.current_project["racks"][rack_index]["items"][item_index]["quantity"]
        new_qty_str = self.simple_input_dialog(self.root, "修改数量", f"请输入新数量 (原: {old_qty})")
        if new_qty_str:
            try:
                new_qty = int(new_qty_str)
                if new_qty <= 0:
                    raise ValueError
                self.project_mgr.update_item_quantity(rack_index, item_index, new_qty)
                self.on_rack_select()
            except:
                messagebox.showerror("错误", "数量必须是正整数")

    def del_rack_detail_item(self):
        """删除机架电气元件明细中选中的直接元件"""
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            messagebox.showwarning("警告", "请先选择一个机架")
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        selection = self.rack_detail_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要删除的元件")
            return
        # 获取选中行的料号
        part_number = self.rack_detail_tree.item(selection[0])["values"][0]
        # 在机架条目中找到对应的元件条目并删除
        items = self.project_mgr.current_project["racks"][rack_index]["items"]
        for i, item in enumerate(items):
            if item["type"] == "element" and item["id"] == part_number:
                del items[i]
                break
        # 刷新显示（会重新加载当前机架的内容和明细）
        self.on_rack_select()

    def edit_rack_detail_qty(self):
        """修改机架电气元件明细中选中的元件数量"""
        rack_sel = self.rack_tree.selection()
        if not rack_sel:
            messagebox.showwarning("警告", "请先选择一个机架")
            return
        rack_index = self.rack_tree.index(rack_sel[0])
        selection = self.rack_detail_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请选择要修改数量的元件")
            return
        # 获取选中行的料号
        part_number = self.rack_detail_tree.item(selection[0])["values"][0]
        items = self.project_mgr.current_project["racks"][rack_index]["items"]
        for item in items:
            if item["type"] == "element" and item["id"] == part_number:
                old_qty = item["quantity"]
                new_qty_str = self.simple_input_dialog(self.root, "修改数量", f"请输入新数量 (原: {old_qty})")
                if new_qty_str:
                    try:
                        new_qty = int(new_qty_str)
                        if new_qty <= 0:
                            raise ValueError
                        item["quantity"] = new_qty
                        self.on_rack_select()  # 刷新界面
                    except:
                        messagebox.showerror("错误", "数量必须是正整数")
                return
        messagebox.showerror("错误", "未找到该元件")

    def new_project(self):
        name = self.simple_input_dialog(self.root, "新建项目", "项目名称:")
        if name:
            self.project_mgr.new_project(name)
            self.project_name_var.set(name)
            self.refresh_rack_display()

    def open_project(self):
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
        if file_path:
            self.project_mgr.load_project(file_path)
            self.project_name_var.set(self.project_mgr.current_project["name"])
            self.refresh_rack_display()

    def save_project(self):
        if not self.project_mgr.file_path:
            file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
            if not file_path:
                return
            self.project_mgr.save_project(file_path)
        else:
            self.project_mgr.save_project()
        messagebox.showinfo("成功", "项目保存成功")

    def manage_component_lib(self):
        dialog = ComponentLibraryDialog(self.root, self.component_db)
        self.root.wait_window(dialog.dialog)
        self.refresh_component_list()

    def manage_electrical_lib(self):
        dialog = ElectricalLibraryDialog(self.root, self.electrical_lib, quantity_mode=False)
        self.root.wait_window(dialog.dialog)
        self.filter_elements()

    def generate_bom(self):
        if not self.project_mgr.current_project["racks"]:
            messagebox.showwarning("警告", "项目中还没有任何机架，请先添加机架和条目")
            return

        bom_window = tk.Toplevel(self.root)
        bom_window.title("电气BOM详情")
        bom_window.geometry("1200x600")
        bom_window.transient(self.root)
        bom_window.grab_set()

        main_frame = ttk.Frame(bom_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text=f"电气BOM - {self.project_mgr.current_project['name']}",
                  font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(0, 10))

        columns = ("rack", "placement", "material", "part_number", "specification", "quantity", "notes")
        tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=25)
        tree.heading("rack", text="机架")
        tree.heading("placement", text="位置")
        tree.heading("material", text="材料名称")
        tree.heading("part_number", text="料号")
        tree.heading("specification", text="规格")
        tree.heading("quantity", text="数量")
        tree.heading("notes", text="备注")
        for col in columns:
            tree.column(col, width=120)
        tree.column("material", width=150)

        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        bom_data = []
        for rack in self.project_mgr.current_project["racks"]:
            rack_name = rack["name"]
            for item in rack["items"]:
                if item["type"] == "element":
                    elem = self.electrical_lib.get_element(item["id"])
                    if elem:
                        final_pn = self.electrical_lib.get_replacement_chain(item["id"]) or item["id"]
                        final_elem = self.electrical_lib.get_element(final_pn) or elem
                        placement_cn = "零散件"
                        tree.insert("", tk.END, values=(
                            rack_name, placement_cn,
                            final_elem["name"], final_pn, final_elem["specification"],
                            item["quantity"], ""
                        ))
                        bom_data.append((rack_name, placement_cn, final_elem["name"], final_pn,
                                         final_elem["specification"], item["quantity"], ""))
                else:
                    comp = self.component_db.get_component(item["id"])
                    if comp:
                        for part in comp["electrical_parts"]:
                            if part["prefab_qty"] > 0:
                                total_prefab = part["prefab_qty"] * item["quantity"]
                                final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                final_elem = self.electrical_lib.get_element(final_pn)
                                material_name = final_elem["name"] if final_elem else part["material"]
                                spec = final_elem["specification"] if final_elem else part["specification"]
                                tree.insert("", tk.END, values=(
                                    rack_name, "预制板", material_name, final_pn, spec, total_prefab, part["notes"]
                                ))
                                bom_data.append((rack_name, "预制板", material_name, final_pn, spec, total_prefab, part["notes"]))
                            if part["spare_qty"] > 0:
                                total_spare = part["spare_qty"] * item["quantity"]
                                final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                final_elem = self.electrical_lib.get_element(final_pn)
                                material_name = final_elem["name"] if final_elem else part["material"]
                                spec = final_elem["specification"] if final_elem else part["specification"]
                                tree.insert("", tk.END, values=(
                                    rack_name, "零散件", material_name, final_pn, spec, total_spare, part["notes"]
                                ))
                                bom_data.append((rack_name, "零散件", material_name, final_pn, spec, total_spare, part["notes"]))

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="导出Excel", command=lambda: self.export_bom_excel(bom_data)).pack()

    def export_bom_excel(self, bom_data):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        try:
            df = pd.DataFrame(bom_data, columns=["机架", "位置", "材料名称", "料号", "规格", "数量", "备注"])
            grouped = df.groupby(["机架", "位置", "料号", "材料名称", "规格", "备注"], as_index=False)["数量"].sum()
            position_order = {"预制板": 0, "零散件": 1}
            grouped["位置排序"] = grouped["位置"].map(position_order)
            grouped = grouped.sort_values(["机架", "位置排序", "料号"]).drop(columns=["位置排序"])
            df_material = grouped[["机架", "位置", "料号", "材料名称", "规格", "数量", "备注"]]

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_material.to_excel(writer, sheet_name="物料明细", index=False)

            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            wb = load_workbook(file_path)
            ws = wb["物料明细"]

            merge_map = {}
            current_key = None
            start_row = 2
            for row in range(2, ws.max_row + 2):
                rack = ws.cell(row=row, column=1).value if row <= ws.max_row else None
                pos = ws.cell(row=row, column=2).value if row <= ws.max_row else None
                key = (rack, pos)
                if key != current_key:
                    if current_key is not None:
                        merge_map[current_key] = (start_row, row - 1)
                    current_key = key
                    start_row = row

            for (rack, pos), (s, e) in merge_map.items():
                if s < e:
                    ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                    ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)
                    ws.cell(row=s, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=s, column=2).alignment = Alignment(horizontal='center', vertical='center')
            wb.save(file_path)

            io_rows = []
            for rack in self.project_mgr.current_project["racks"]:
                rack_name = rack["name"]
                input_names = []
                output_names = []
                for item in rack["items"]:
                    if item["type"] != "component":
                        continue
                    comp = self.component_db.get_component(item["id"])
                    if not comp or not comp.get("io_points"):
                        continue
                    io_points = comp["io_points"]
                    for instance in range(1, item["quantity"] + 1):
                        for name in io_points.get("inputs", []):
                            full_name = f"{comp['name']}{instance}{name}"
                            input_names.append(full_name)
                        for name in io_points.get("outputs", []):
                            full_name = f"{comp['name']}{instance}{name}"
                            output_names.append(full_name)
                for idx, name in enumerate(input_names, start=1):
                    io_rows.append({"机架": rack_name, "点类型": "输入", "序号": idx, "点名称": name})
                for idx, name in enumerate(output_names, start=1):
                    io_rows.append({"机架": rack_name, "点类型": "输出", "序号": idx, "点名称": name})
            df_io = pd.DataFrame(io_rows) if io_rows else pd.DataFrame(columns=["机架", "点类型", "序号", "点名称"])
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_io.to_excel(writer, sheet_name="IO分配表", index=False)

            messagebox.showinfo("成功", f"BOM已导出到: {file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def export_excel(self):
        if self.project_mgr.current_project["racks"]:
            bom_data = []
            for rack in self.project_mgr.current_project["racks"]:
                rack_name = rack["name"]
                for item in rack["items"]:
                    if item["type"] == "element":
                        elem = self.electrical_lib.get_element(item["id"])
                        if elem:
                            final_pn = self.electrical_lib.get_replacement_chain(item["id"]) or item["id"]
                            final_elem = self.electrical_lib.get_element(final_pn) or elem
                            bom_data.append((rack_name, "零散件", final_elem["name"], final_pn,
                                             final_elem["specification"], item["quantity"], ""))
                    else:
                        comp = self.component_db.get_component(item["id"])
                        if comp:
                            for part in comp["electrical_parts"]:
                                if part["prefab_qty"] > 0:
                                    total = part["prefab_qty"] * item["quantity"]
                                    final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                    final_elem = self.electrical_lib.get_element(final_pn)
                                    name = final_elem["name"] if final_elem else part["material"]
                                    spec = final_elem["specification"] if final_elem else part["specification"]
                                    bom_data.append((rack_name, "预制板", name, final_pn, spec, total, part["notes"]))
                                if part["spare_qty"] > 0:
                                    total = part["spare_qty"] * item["quantity"]
                                    final_pn = self.electrical_lib.get_replacement_chain(part["part_number"]) or part["part_number"]
                                    final_elem = self.electrical_lib.get_element(final_pn)
                                    name = final_elem["name"] if final_elem else part["material"]
                                    spec = final_elem["specification"] if final_elem else part["specification"]
                                    bom_data.append((rack_name, "零散件", name, final_pn, spec, total, part["notes"]))
            self.export_bom_excel(bom_data)
        else:
            messagebox.showwarning("警告", "项目为空，无法导出")

    def simple_input_dialog(self, parent, title, prompt):
        result = [None]
        dialog = tk.Toplevel(parent)
        dialog.title(title)
        dialog.geometry("300x120")
        dialog.transient(parent)
        dialog.grab_set()
        dialog.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() // 2) - (dialog.winfo_width() // 2)
        y = parent.winfo_y() + (parent.winfo_height() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=prompt).pack(anchor=tk.W, pady=(0, 5))
        entry = ttk.Entry(frame, width=30)
        entry.pack(fill=tk.X, pady=(0, 10))
        entry.focus()
        def ok():
            result[0] = entry.get().strip()
            dialog.destroy()
        def cancel():
            result[0] = None
            dialog.destroy()
        btn_frame = ttk.Frame(frame)
        btn_frame.pack()
        ttk.Button(btn_frame, text="确定", command=ok).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="取消", command=cancel).pack(side=tk.LEFT)
        entry.bind('<Return>', lambda e: ok())
        dialog.wait_window()
        return result[0]

    def show_component_details_popup(self, event):
        """双击机架内容中的组件，弹出窗口显示该组件的电气元件明细"""
        selection = self.rack_items_tree.selection()
        if not selection:
            return
        item = self.rack_items_tree.item(selection[0])
        item_type = item["values"][0]
        if item_type != "组件":
            return
        component_name = item["values"][1]
        comp = self.component_db.get_component(component_name)
        if not comp:
            messagebox.showerror("错误", f"未找到组件 '{component_name}'")
            return

        popup = tk.Toplevel(self.root)
        popup.title(f"组件电气元件清单 - {component_name}")
        popup.geometry("800x400")
        popup.transient(self.root)
        popup.grab_set()

        # 居中显示
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (popup.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (popup.winfo_height() // 2)
        popup.geometry(f"+{x}+{y}")

        # 表格
        columns = ("material", "prefab_qty", "spare_qty", "part_number", "specification", "detailed_spec", "notes")
        tree = ttk.Treeview(popup, columns=columns, show="headings")
        tree.heading("material", text="材料名称")
        tree.heading("prefab_qty", text="预制板数量")
        tree.heading("spare_qty", text="零散件数量")
        tree.heading("part_number", text="料号")
        tree.heading("specification", text="规格型号")
        tree.heading("detailed_spec", text="详细规格")
        tree.heading("notes", text="备注")
        for col in columns:
            tree.column(col, width=100)
        tree.column("material", width=120)

        scrollbar = ttk.Scrollbar(popup, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 填充数据
        for part in comp["electrical_parts"]:
            tree.insert("", tk.END, values=(
                part["material"],
                part["prefab_qty"],
                part["spare_qty"],
                part["part_number"],
                part["specification"],
                part["detailed_spec"],
                part["notes"]
            ))

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="关闭", command=popup.destroy).pack()